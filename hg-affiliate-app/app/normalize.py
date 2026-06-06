"""사업자명(거래처명) 정규화 (절대 준수 c).

처리 순서 (참고 의사코드 반영):
  1) "특관자 상호 정리" 파일의 들여쓰기(indent) 구조로 1차 매핑
     - indent 0 = 정규(대표) 법인명, indent>=1 = 이표기(별칭)
  2) 미매칭 사업자명만 ChatGPT API 로 대표 법인명 매핑
  3) AI 실패/키 없음 시 휴리스틱 폴백(정규화 키 기반 근사 매칭)

비용·레이트리밋 대비: 배치 호출 + 캐시 + 재시도.
API 키 값은 화면·로그에 노출하지 않는다.
"""

from __future__ import annotations

import json
import re
import time
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .config import Settings
from .domain.columns import is_aggregate_label


@dataclass
class NormalizationResult:
    # 이표기 -> 정규명
    mapping: dict[str, str] = field(default_factory=dict)
    # 정규명 집합
    canonical: set[str] = field(default_factory=set)
    # AI 로도 해소 못한 미매칭 (오류목록에 기록)
    unmatched: list[str] = field(default_factory=list)
    # 사용한 방법 요약 (수치 미포함)
    method_notes: list[str] = field(default_factory=list)


def _norm_key(name: str) -> str:
    """비교용 정규화 키: 공백/괄호/법인격/지점 표기 제거 후 소문자."""
    s = str(name).strip()
    s = re.sub(r"[\s\u3000]", "", s)
    s = s.replace("㈜", "").replace("(주)", "").replace("주식회사", "")
    s = re.sub(r"(지점|점|본사|영업소|사업부)$", "", s)
    s = re.sub(r"[()\[\]·.,/\\-]", "", s)
    return s.lower()


def seed_from_name_pivot(path: Path | None) -> dict[str, str]:
    """특관자 상호 정리 파일의 들여쓰기 구조로 1차 매핑을 만든다.

    openpyxl 로 A열을 순회하며 alignment.indent 를 사용.
    """
    norm: dict[str, str] = {}
    if path is None or not path.exists() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return norm
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
    except Exception:  # noqa: BLE001 - xls/csv 등은 시드 건너뜀
        return norm

    cur: str | None = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            val = cell.value
            if val is None or not str(val).strip():
                continue
            name = str(val).strip()
            if is_aggregate_label(name):  # 총합계/행 레이블 등 피벗 산출물 제외
                continue
            indent = 0
            try:
                indent = int(cell.alignment.indent or 0)
            except Exception:  # noqa: BLE001
                indent = 0
            if indent == 0:
                cur = name
                norm[name] = name
            elif cur:
                norm[name] = cur
    wb.close()
    return norm


def _heuristic_fallback(unmatched: list[str], norm: dict[str, str]) -> dict[str, str]:
    """정규화 키가 같은 기존 정규명에 근사 매칭. 없으면 자기 자신을 정규명으로."""
    key_to_canon: dict[str, str] = {}
    for alias, canon in norm.items():
        key_to_canon.setdefault(_norm_key(canon), canon)

    out: dict[str, str] = {}
    for name in unmatched:
        k = _norm_key(name)
        if k in key_to_canon:
            out[name] = key_to_canon[k]
        else:
            # 같은 미매칭 그룹 내에서도 키가 같으면 첫 표기를 대표로
            key_to_canon.setdefault(k, name)
            out[name] = key_to_canon[k]
    return out


def _openai_map(unmatched: list[str], seed: dict[str, str], settings: Settings) -> dict[str, str]:
    """ChatGPT API 로 대표 법인명 매핑. 배치 + 재시도."""
    from openai import OpenAI  # 지연 임포트

    client = OpenAI(api_key=settings.openai_api_key)
    canon_examples = sorted(set(seed.values()))[:200]
    result: dict[str, str] = {}

    batch_size = 60
    for i in range(0, len(unmatched), batch_size):
        batch = unmatched[i : i + batch_size]
        prompt = _build_prompt(batch, canon_examples)
        text = _chat_with_retry(client, settings.openai_model, prompt)
        if text:
            result.update(_parse_mapping(text, batch))
    return result


def _build_prompt(names: list[str], canon_examples: list[str]) -> str:
    return (
        "다음은 회계 원장에 등장한 거래처 '사업자명' 표기 목록입니다. "
        "약어·띄어쓰기·지점표기·구사명 등으로 같은 회사가 다르게 적혀 있을 수 있습니다. "
        "각 표기를 대표(정규) 법인명으로 매핑하세요. "
        "가능하면 아래 기준 법인명 중 하나로 매핑하고, 없으면 가장 정제된 정규 법인명을 새로 제시하세요. "
        "주의: 'HLB(에이치엘비)' 와 'HLB글로벌(에이치엘비글로벌)' 과 'HLB생명과학' 은 서로 다른 회사입니다.\n\n"
        f"기준 법인명(예시): {json.dumps(canon_examples, ensure_ascii=False)}\n\n"
        f"매핑할 표기 목록: {json.dumps(names, ensure_ascii=False)}\n\n"
        '반드시 JSON 객체만 출력하세요. 형식: {"표기": "정규법인명", ...}'
    )


def _chat_with_retry(client, model: str, prompt: str, retries: int = 3) -> str | None:
    for attempt in range(retries):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "너는 한국 회계 거래처명 정규화 전문가다. JSON만 출력한다."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0,
                response_format={"type": "json_object"},
            )
            return resp.choices[0].message.content
        except Exception:  # noqa: BLE001 - 레이트리밋/일시오류 재시도
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
    return None


def _parse_mapping(text: str, batch: list[str]) -> dict[str, str]:
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if not m:
            return {}
        try:
            data = json.loads(m.group(0))
        except json.JSONDecodeError:
            return {}
    out: dict[str, str] = {}
    for k, v in data.items():
        if isinstance(v, str) and v.strip():
            out[str(k)] = v.strip()
    return out


def normalize_names(
    distinct_names: list[str],
    name_pivot_path: Path | None,
    settings: Settings,
) -> NormalizationResult:
    """전체 정규화 수행."""
    result = NormalizationResult()

    # 1) indent 시드
    norm = seed_from_name_pivot(name_pivot_path)
    if norm:
        result.method_notes.append("특관자 상호 정리(들여쓰기) 1차 매핑 적용")

    # 시드의 정규명 집합 (특관자 후보)
    canonical = set(norm.values())

    # 2) 미매칭 추출
    unmatched = [n for n in distinct_names if n and n not in norm]

    if unmatched:
        if settings.has_openai_key:
            try:
                ai_map = _openai_map(unmatched, norm, settings)
                norm.update(ai_map)
                result.method_notes.append("ChatGPT API 정규화 적용")
            except Exception:  # noqa: BLE001 - 폴백으로 전환
                result.method_notes.append("ChatGPT 호출 실패 → 휴리스틱 폴백")
        else:
            result.method_notes.append("OPENAI_API_KEY 미설정 → 휴리스틱 폴백")

        still_unmatched = [n for n in unmatched if n not in norm]
        if still_unmatched:
            norm.update(_heuristic_fallback(still_unmatched, norm))

    # 피벗/집계 산출물이 매핑·정규명에 섞였으면 최종 제거 (방어)
    norm = {
        alias: canon
        for alias, canon in norm.items()
        if not is_aggregate_label(alias) and not is_aggregate_label(canon)
    }
    canonical = {c for c in (canonical or set(norm.values())) if not is_aggregate_label(c)}

    result.mapping = norm
    final_canon = canonical if canonical else set(norm.values())
    result.canonical = final_canon

    # 검토 대상(unmatched) 산정:
    #  - 특관자 목록(시드)이 있으면: 목록에 없는 '일반 거래처'는 특관자가 아니므로
    #    오류목록에 올리지 않는다(원장에는 수백 개의 일반 거래처가 있어 노이즈가 됨).
    #    다만 정규(대표) 법인명과 표기가 부분적으로 겹치는데 매칭되지 못한 건은
    #    '특관자 별칭 누락 의심'으로 검토 대상에 남긴다.
    #  - 시드가 없으면: 종전처럼 정규명에 들지 못한 표기를 검토 대상으로 한다.
    if norm and canonical:
        canon_keys = {_norm_key(c) for c in final_canon if _norm_key(c)}
        review: list[str] = []
        for n in distinct_names:
            if not n or norm.get(n) in final_canon:
                continue
            k = _norm_key(n)
            if len(k) < 3:
                continue
            if any(len(ck) >= 3 and (k in ck or ck in k) for ck in canon_keys):
                review.append(n)
        result.unmatched = review
    else:
        result.unmatched = [n for n in distinct_names if n and norm.get(n) not in final_canon]
    return result
