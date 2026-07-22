"""당기 특관자 명세서 생성 (출력물 1).

시트 구성: 특관자 / 39.1 / 39.2 / _집계(데이터)
- _집계: 정규명별 38.x 산출값(데이터 소스)
- 특관자: SUMIF 로 _집계를 자동 합산(사업보고서 게재용). 수식만 사용해 재현성 확보.
- 39.2: 정규명 x 항목 피벗.

수식은 동일 워크북 내 _집계 시트만 참조하므로 LibreOffice 재계산 시 오류가 없다.
값/수치는 화면·로그로 내보내지 않으며 엑셀 파일에만 기록한다.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .aggregate import AggregateResult
from .period_extract import Period
from .statement_detail import (
    StatementDetailRow,
    detail_balance_formula,
    detail_group_key,
    write_detail_cell,
)
from .statement_template import fill_statement_template, write_period_label

# 고정 템플릿(정답 양식). 존재하면 이 양식으로 산출, 없으면 단순표 폴백.
TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "assets" / "statement_template.xlsx"

# (표시 헤더, Aggregate 속성) — 38.5 경영진보상은 공란이라 값 컬럼 없음
COLUMNS: list[tuple[str, str]] = [
    ("매출", "sales"),
    ("매출등 기타", "sales_other"),
    ("매입", "purchase"),
    ("매입등 기타", "purchase_other"),
    ("매출채권", "ar"),
    ("대여금", "loan"),
    ("기타채권", "other_ar"),
    ("투자전환사채", "cb_invest"),
    ("기타채무", "other_ap"),
    ("발행전환사채", "cb_issued"),
    ("매입채무", "ap"),
    ("대손충당금", "allowance_end"),
    ("대손상각비", "allowance_expense"),
    ("자금대여", "fund_lending"),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_NUMFMT = "#,##0"


def _style_header(cell) -> None:
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER
    cell.border = _BORDER


def build_statement(
    result: AggregateResult,
    period: Period,
    detail_rows: Sequence[StatementDetailRow],
) -> tuple[Workbook, list[str]]:
    """당기 특관자 명세서 워크북을 만든다.

    Returns:
        (workbook, unmatched) — 고정 템플릿 사용 시 템플릿 거래처에 매칭되지 못한
        정규명 목록(수기 추가 검토). 폴백 경로에선 빈 목록.
    """
    if TEMPLATE_PATH.exists():
        return fill_statement_template(TEMPLATE_PATH, result, period, detail_rows)
    return _build_legacy(result, period, detail_rows), []


def _build_legacy(
    result: AggregateResult,
    period: Period,
    detail_rows: Sequence[StatementDetailRow],
) -> Workbook:
    wb = Workbook()

    data_ws = wb.active
    data_ws.title = "_집계"
    _write_data_sheet(data_ws, result)

    main_ws = wb.create_sheet("특관자")
    _write_main_sheet(main_ws, result, data_ws.title)

    pivot_ws = wb.create_sheet("39.2")
    _write_pivot_sheet(pivot_ws, result)

    note_ws = wb.create_sheet("39.1")
    write_period_label(note_ws, period)
    headers = [
        "매입매출", "자금거래", "채권채무", "수익.비용", "구분계정과목",
        "계정코드", "계정과목", "날짜", "적요", "거래처코드", "거래처명",
        "회사명(본지점합산)", "차변", "대변", "잔액",
    ]
    for column, header in enumerate(headers, start=2):
        _style_header(note_ws.cell(row=15, column=column, value=header))
    group_first_row = 16
    previous_group_key: tuple[str, str, str] | None = None
    for row_number, detail in enumerate(detail_rows, start=16):
        current_group_key = detail_group_key(detail)
        if current_group_key != previous_group_key:
            group_first_row = row_number
        previous_group_key = current_group_key
        for column, value in enumerate(detail.as_excel_row(), start=2):
            cell = note_ws.cell(row=row_number, column=column)
            if column == 16 and detail.balance is not None:
                cell.value = detail_balance_formula(
                    detail, group_first_row, row_number
                )
            else:
                write_detail_cell(cell, value)
    # _집계 시트는 데이터 소스이므로 숨김
    data_ws.sheet_state = "hidden"
    wb.active = wb.sheetnames.index("특관자")
    return wb


def _write_data_sheet(ws, result: AggregateResult) -> None:
    headers = ["정규명"] + [h for h, _ in COLUMNS]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        _style_header(c)
    names = sorted(result.by_canonical.keys())
    for r, name in enumerate(names, start=2):
        agg = result.by_canonical[name]
        ws.cell(row=r, column=1, value=name).border = _BORDER
        for col, (_, attr) in enumerate(COLUMNS, start=2):
            cell = ws.cell(row=r, column=col, value=round(getattr(agg, attr), 0))
            cell.number_format = _NUMFMT
            cell.border = _BORDER
    ws.cell(row=1, column=1).comment = None


def _write_main_sheet(ws, result: AggregateResult, data_title: str) -> None:
    ws["A1"] = "HLB글로벌(주) 당기 특수관계자 명세서"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS) + 1)

    header_row = 3
    headers = ["정규(대표) 법인명"] + [h for h, _ in COLUMNS]
    for col, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=header_row, column=col, value=h))

    names = sorted(result.by_canonical.keys())
    n = len(names)
    first_data = header_row + 1
    for i, name in enumerate(names):
        r = first_data + i
        ws.cell(row=r, column=1, value=name).border = _BORDER
        # _집계 시트의 같은 정규명 행을 SUMIF 로 자동 합산
        for col in range(2, len(COLUMNS) + 2):
            col_letter = get_column_letter(col)
            crit_range = f"'{data_title}'!$A:$A"
            sum_range = f"'{data_title}'!${col_letter}:${col_letter}"
            formula = f"=SUMIF({crit_range},$A{r},{sum_range})"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = _NUMFMT
            cell.border = _BORDER

    # 합계 행
    if n:
        total_r = first_data + n
        ws.cell(row=total_r, column=1, value="합계").font = Font(bold=True)
        for col in range(2, len(COLUMNS) + 2):
            col_letter = get_column_letter(col)
            rng = f"{col_letter}{first_data}:{col_letter}{first_data + n - 1}"
            cell = ws.cell(row=total_r, column=col, value=f"=SUM({rng})")
            cell.number_format = _NUMFMT
            cell.font = Font(bold=True)
            cell.border = _BORDER

    ws.column_dimensions["A"].width = 28
    for col in range(2, len(COLUMNS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _write_pivot_sheet(ws, result: AggregateResult) -> None:
    ws["A1"] = "39.2 특수관계자별 거래/잔액 피벗"
    ws["A1"].font = Font(bold=True)
    header_row = 3
    headers = ["정규명"] + [h for h, _ in COLUMNS]
    for col, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=header_row, column=col, value=h))
    for i, name in enumerate(sorted(result.by_canonical.keys())):
        r = header_row + 1 + i
        agg = result.by_canonical[name]
        ws.cell(row=r, column=1, value=name).border = _BORDER
        for col, (_, attr) in enumerate(COLUMNS, start=2):
            cell = ws.cell(row=r, column=col, value=round(getattr(agg, attr), 0))
            cell.number_format = _NUMFMT
            cell.border = _BORDER
    ws.column_dimensions["A"].width = 28
