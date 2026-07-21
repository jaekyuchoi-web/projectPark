"""정답 양식(고정 템플릿) 기반 당기 특관자 명세서 생성.

전략(핵심):
- `app/assets/statement_template.xlsx` = 사업보고서 게재용 '특관자' 양식 + '39.1'/'39.2'.
- '특관자' 시트는 전부 `=SUMIF('39.2'!$C$3:$C$44, ...)` / `'39.2'!$C$49:$C$100` 수식이라,
  실제로는 **'39.2' 시트의 값 컬럼만 채우면** 보고서 수치가 자동 계산된다.
    · 블록1(거래, 3~44행):  E=매출  F=기타수익  G=매입  H=기타비용
    · 블록2(채권채무, 50~92행): E=매출채권 F=대여금 G=기타채권 [H=투자전환사채] I=기타채무 [J=발행전환사채]
      └ 단, 투자/발행 전환사채(H·J)는 '회차별 보유자 확인'이 필요한 수기 영역이라
        자동 주입하지 않고 템플릿 값을 보존한다(정답 파일에서도 공란).
- 거래처명(D열)을 '정규명'으로 매칭해 우리 집계값을 주입한다(정규화 키 비교).
- 템플릿에는 외부 워크북 참조('[29]39.1 (2)' 등 351개)와 기존 `#REF!`(4개)가 있어,
  LibreOffice 재계산 시 오류가 된다. 우리가 덮어쓰지 않는 외부/`#REF!` 수식은 **캐시값(또는 0)**
  으로 중화한다(로컬 '39.2'/'특관자' SUMIF 수식은 보존).
- 전기말·38.3~38.5·사채명세 등 수기/전기 영역은 템플릿 값을 그대로 보존한다(분기별 수동 수정 전제).
  · "전기" = 전년도 온기(연간 결산 = 4분기). 같은 해 직전분기 자료는 참고할 수 있으나
    전기가 아니다. 양식의 '전반기' 등 참고 블록은 전기말 비교열과 혼동하지 말 것.

값/수치는 화면·로그로 내보내지 않으며 엑셀 파일에만 기록한다.
"""

from __future__ import annotations

import re
from collections.abc import Sequence
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formula import Tokenizer

from ..normalize import _norm_key
from .aggregate import Aggregate, AggregateResult
from .period_extract import Period
from .statement_detail import StatementDetailRow, write_detail_cell

# 39.2 블록 좌표(헤더 검증 포함)
_B1_FIRST, _B1_LAST = 3, 44       # 블록1: 거래(38.1)
_B2_HEADER = 49                   # 블록2 헤더행
_B2_FIRST, _B2_LAST = 50, 92      # 블록2: 채권채무(38.2)

# (39.2 컬럼 인덱스, Aggregate 속성) — 블록1 거래
_BLOCK1_COLS: list[tuple[int, str]] = [
    (5, "sales"),          # E 매출
    (6, "sales_other"),    # F 기타수익
    (7, "purchase"),       # G 매입
    (8, "purchase_other"), # H 기타비용
]
# (39.2 컬럼 인덱스, Aggregate 속성) — 블록2 채권채무
# 주의: 투자전환사채(H,8)·발행전환사채(J,10)는 의도적으로 제외한다.
#   - 발행/투자 전환사채는 '회차별 보유자(특관자)' 확인이 필요한 수기/검증 영역이라
#     잔액명세서에서 거래처별로 자동 산출할 수 없다(정답 파일에서도 해당 칸은 공란).
#   - 따라서 38.3~38.5·사채명세와 동일하게 '템플릿 값 보존'으로 두고,
#     오류목록('전환사채' 검토 항목)으로만 안내한다. (aggregate.aggregate_balance 참조)
_BLOCK2_COLS: list[tuple[int, str]] = [
    (5, "ar"),         # E 매출채권
    (6, "loan"),       # F 대여금
    (7, "other_ar"),   # G 기타채권
    (9, "other_ap"),   # I 기타채무
]

_EXT_RE = re.compile(r"\[\d+\]")  # 외부 워크북 링크 인덱스
# 셀에 그대로 박힌 엑셀 오류 토큰(수식이 아닌 문자열 값으로 저장된 경우 포함)
_ERR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

_DETAIL_FIRST_ROW = 16
_DETAIL_FIRST_COL = 2
_DETAIL_LAST_COL = 16
_DETAIL_TOKEN_RANGE_RE = re.compile(
    r"(?P<start_col>\$?[A-Za-z]{1,3})(?P<start_row>\$?\d+):"
    r"(?P<end_col>\$?[A-Za-z]{1,3})(?P<end_row>\$?\d+)"
)


class StatementTemplateError(ValueError):
    """The statement template cannot be updated safely."""


def _parse_detail_range_token(value: str, sheet_title: str):
    qualifier = ""
    reference = value
    if "!" in value:
        sheet, reference = value.rsplit("!", 1)
        qualifier = f"{sheet}!"
        if sheet.startswith("'") and sheet.endswith("'"):
            sheet = sheet[1:-1].replace("''", "'")
        if sheet.casefold() != sheet_title.casefold():
            return None

    match = _DETAIL_TOKEN_RANGE_RE.fullmatch(reference)
    if match is None:
        return None
    start_col = match.group("start_col").replace("$", "").upper()
    end_col = match.group("end_col").replace("$", "").upper()
    start_row = int(match.group("start_row").replace("$", ""))
    if (
        start_col != end_col
        or start_col not in tuple("BCDEFGHIJKLMNOP")
        or start_row != 16
    ):
        return None
    return qualifier, reference, match


def _rewrite_detail_range_token(value: str, sheet_title: str, last_row: int) -> str | None:
    parsed = _parse_detail_range_token(value, sheet_title)
    if parsed is None:
        return None
    qualifier, reference, match = parsed

    end_row = match.group("end_row")
    end_marker = "$" if end_row.startswith("$") else ""
    rewritten = (
        reference[: match.start("end_row")]
        + end_marker
        + str(last_row)
        + reference[match.end("end_row") :]
    )
    return qualifier + rewritten


def _rewrite_detail_formula_ranges(ws, last_row: int) -> None:
    recognized = 0
    try:
        for row in ws.iter_rows(min_row=1, max_row=_DETAIL_FIRST_ROW - 1):
            for cell in row:
                formula = cell.value
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                tokenizer = Tokenizer(formula)
                changed = False
                for token in tokenizer.items:
                    if token.type != "OPERAND" or token.subtype != "RANGE":
                        continue
                    rewritten = _rewrite_detail_range_token(
                        token.value, ws.title, last_row
                    )
                    if rewritten is None:
                        continue
                    token.value = rewritten
                    verified = _parse_detail_range_token(token.value, ws.title)
                    if verified is None:
                        raise StatementTemplateError
                    endpoint = int(
                        verified[2].group("end_row").replace("$", "")
                    )
                    if endpoint != last_row:
                        raise StatementTemplateError
                    recognized += 1
                    changed = True
                if changed:
                    cell.value = tokenizer.render()
    except Exception as exc:
        raise StatementTemplateError(
            "39.1 요약 수식을 안전하게 갱신하지 못했습니다."
        ) from exc
    if recognized == 0:
        raise StatementTemplateError(
            "39.1 요약 수식을 안전하게 갱신하지 못했습니다."
        )


def _write_statement_detail(
    ws,
    period: Period,
    rows: Sequence[StatementDetailRow],
) -> None:
    style_source = [
        ws.cell(row=_DETAIL_FIRST_ROW, column=column)
        for column in range(_DETAIL_FIRST_COL, _DETAIL_LAST_COL + 1)
    ]
    source_height = ws.row_dimensions[_DETAIL_FIRST_ROW].height
    clear_through = max(ws.max_row, _DETAIL_FIRST_ROW)
    for row in ws.iter_rows(
        min_row=_DETAIL_FIRST_ROW,
        max_row=clear_through,
        min_col=_DETAIL_FIRST_COL,
        max_col=_DETAIL_LAST_COL,
    ):
        for cell in row:
            cell.value = None
            cell.comment = None
            cell.hyperlink = None

    for offset, detail in enumerate(rows):
        target_row = _DETAIL_FIRST_ROW + offset
        if source_height is not None:
            ws.row_dimensions[target_row].height = source_height
        for index, value in enumerate(detail.as_excel_row()):
            column = _DETAIL_FIRST_COL + index
            target = ws.cell(row=target_row, column=column)
            write_detail_cell(target, value)
            source = style_source[index]
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)

    last_row = _DETAIL_FIRST_ROW + max(len(rows), 1) - 1
    _rewrite_detail_formula_ranges(ws, last_row)
    ws["A1"] = period.label
    label_font = copy(ws["A1"].font)
    label_font.bold = True
    ws["A1"].font = label_font


def _has_data(agg: Aggregate) -> bool:
    # 자동 주입 대상 버킷만 본다. 전환사채(cb_invest/cb_issued)는 수기 영역이라
    # 여기서 제외(전환사채만 있는 법인을 '템플릿 미반영'으로 오보고하지 않도록).
    return any(
        abs(getattr(agg, attr, 0.0) or 0.0) > 0.0
        for attr in (
            "sales", "sales_other", "purchase", "purchase_other",
            "ar", "loan", "other_ar", "other_ap",
        )
    )


def fill_statement_template(
    template_path: Path,
    result: AggregateResult,
    period: Period,
    detail_rows: Sequence[StatementDetailRow],
) -> tuple[Workbook, list[str]]:
    """고정 템플릿에 당기 집계값을 주입한다.

    Returns:
        (workbook, unmatched) — unmatched: 데이터가 있으나 템플릿 거래처 행에
        매칭되지 못한 정규명 목록(수기 추가 검토 대상).
    """
    wb = load_workbook(template_path, data_only=False)
    wb_cache = load_workbook(template_path, data_only=True)
    _write_statement_detail(wb["39.1"], period, detail_rows)

    # 정규명(정규화 키) → Aggregate
    by_key: dict[str, Aggregate] = {}
    for canon, agg in result.by_canonical.items():
        k = _norm_key(canon)
        if k:
            by_key[k] = agg
    matched_keys: set[str] = set()

    n2 = wb["39.2"]
    n2c = wb_cache["39.2"]

    def cached(ws_cache, r: int, c: int):
        v = ws_cache.cell(row=r, column=c).value
        return v if isinstance(v, (int, float)) else None

    # ── 블록1: 거래(38.1) ───────────────────────────────────────────────
    for r in range(_B1_FIRST, _B1_LAST + 1):
        dname = n2.cell(row=r, column=4).value  # D 거래처명
        key = _norm_key(dname) if dname else ""
        agg = by_key.get(key)
        if key and agg is not None:
            matched_keys.add(key)
        for col, attr in _BLOCK1_COLS:
            val = float(round(getattr(agg, attr), 0)) if agg is not None else 0.0
            n2.cell(row=r, column=col, value=val)
        # L:R 상세(외부참조)는 캐시값으로 보존(없으면 0)
        for col in range(12, 19):  # L..R
            cell = n2.cell(row=r, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = cached(n2c, r, col) or 0

    # ── 블록2: 채권채무(38.2) ───────────────────────────────────────────
    for r in range(_B2_FIRST, _B2_LAST + 1):
        dname = n2.cell(row=r, column=4).value
        key = _norm_key(dname) if dname else ""
        agg = by_key.get(key)
        if key and agg is not None:
            matched_keys.add(key)
        for col, attr in _BLOCK2_COLS:
            val = float(round(getattr(agg, attr), 0)) if agg is not None else 0.0
            n2.cell(row=r, column=col, value=val)
        # K(지분취득) 등 외부참조 컬럼은 캐시값 보존
        cell_k = n2.cell(row=r, column=11)
        if isinstance(cell_k.value, str) and cell_k.value.startswith("="):
            cell_k.value = cached(n2c, r, 11) or 0

    # ── 외부참조('[n]')·#REF! 수식 전면 중화(덮어쓰지 않은 잔여분) ─────────
    for ws in wb.worksheets:
        wsc = wb_cache[ws.title]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str):
                    continue
                if v.startswith("="):
                    # 외부참조/#REF! 수식 → 캐시값(또는 공란)으로 중화
                    if _EXT_RE.search(v) or any(t in v for t in _ERR_TOKENS):
                        c.value = cached(wsc, c.row, c.column)
                elif v in _ERR_TOKENS:
                    # 문자열 값으로 박힌 오류 토큰(예: 상세설명 셀) → 공란 처리
                    c.value = None

    # ── 외부 링크·명명된 범위 정의 제거 ─────────────────────────────────
    # 수식을 중화해도 워크북에 외부 링크 파트(/xl/externalLinks/*.xml)가 남아 있으면
    # Excel 이 "외부 수식 참조 복구"를 시도한다. 모든 참조를 캐시값으로 바꿨으므로
    # 외부 링크 정의를 통째로 제거한다.
    try:
        wb._external_links = []  # noqa: SLF001 — openpyxl 내부 구조 직접 정리
    except Exception:  # noqa: BLE001
        pass
    # 정답 템플릿에는 세무/외부통합용 명명된 범위가 5만+개(#REF!/외부참조 포함) 남아
    # 있어 Excel 이 "명명된 범위 제거" 복구를 띄운다. 특관자/39.x 시트 수식은 셀
    # 참조만 쓰므로 전부 제거한다(저장 시 시트 자동필터용 _FilterDatabase 만 재생성).
    wb.defined_names.clear()

    # ── 매칭 실패(데이터 보유) 정규명 목록 ──────────────────────────────
    unmatched = [
        canon for canon, agg in result.by_canonical.items()
        if _has_data(agg) and _norm_key(canon) not in matched_keys
    ]

    if "특관자" in wb.sheetnames:
        wb.active = wb.sheetnames.index("특관자")
    return wb, unmatched
