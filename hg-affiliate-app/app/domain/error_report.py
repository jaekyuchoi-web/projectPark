"""오류 목록 출력물 생성 (출력물 3)."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .errors import ErrorLog

_HEADER_FILL = PatternFill("solid", fgColor="7C3A00")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="center", wrap_text=True)


def build_error_report(log: ErrorLog) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "오류목록"

    headers = ["구분", "대상", "내용", "권고조치"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    if not log.items:
        ws.cell(row=2, column=1, value="없음")
        ws.cell(row=2, column=3, value="검토가 필요한 항목이 없습니다.")

    for i, item in enumerate(log.items):
        r = 2 + i
        for col, v in enumerate([item.category, item.target, item.content, item.action], start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = _BORDER
            c.alignment = _WRAP

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 40
    return wb
