"""지배구조 출력물 생성 (출력물 2).

정규 법인명별 [구분, HLB글로벌과의 관계, 분개장 이표기(별칭), 비고] 표.
구분(종속/관계/기타특수관계자/주요경영진)은 자동 판정이 어려우므로
기본값을 두고 비고에 검토 요청을 남긴다.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .columns import is_aggregate_label

_HEADER_FILL = PatternFill("solid", fgColor="375623")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="center", wrap_text=True)


def build_governance(mapping: dict[str, str], canonical: set[str]) -> Workbook:
    """mapping: 이표기->정규명, canonical: 정규명 집합."""
    wb = Workbook()
    ws = wb.active
    ws.title = "지배구조"

    ws["A1"] = "HLB글로벌(주) 특수관계자 지배구조"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    headers = ["정규(대표) 법인명", "구분", "HLB글로벌과의 관계", "분개장 이표기(별칭)", "비고"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    # 총합계/행 레이블 등 비법인 라벨은 최종 출력에서 제외 (방어)
    canonical = {c for c in canonical if not is_aggregate_label(c)}

    # 정규명별 별칭 모으기 (별칭도 피벗 산출물이면 제외)
    aliases: dict[str, list[str]] = {}
    for alias, canon in mapping.items():
        if canon in canonical and alias != canon and not is_aggregate_label(alias):
            aliases.setdefault(canon, []).append(alias)

    names = sorted(canonical)
    for i, name in enumerate(names):
        r = 4 + i
        alias_text = ", ".join(sorted(set(aliases.get(name, []))))
        row_vals = [name, "기타특수관계자", "특수관계자", alias_text, "구분/관계는 검토 후 확정 필요"]
        for col, v in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = _BORDER
            c.alignment = _WRAP

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 28
    return wb
