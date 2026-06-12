"""출력 파일 안전성 검증 (절대 준수 a).

생성한 출력 xlsx 를 다시 열어 검증한다.
  1) openpyxl 재오픈 (구조 무결성)
  2) LibreOffice headless 재계산 (soffice --convert-to xlsx) — 가능 시
  3) 재계산본 전 셀 스캔 → 수식 오류 0 확인
     (#REF! #VALUE! #DIV/0! #N/A #NAME? #NULL! #NUM!)

오류가 1건이라도 있으면 실패로 처리하여 다운로드 차단.
LibreOffice 미설치 시 openpyxl 재오픈 + 수식 정적 점검으로 대체한다.
"""

from __future__ import annotations

import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


@dataclass
class CheckResult:
    ok: bool
    reason: str
    recalculated: bool  # LibreOffice 재계산 수행 여부


def _find_soffice() -> str | None:
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    # macOS 기본 설치 경로
    mac = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    if mac.exists():
        return str(mac)
    return None


def _scan_workbook_for_errors(path: Path) -> str | None:
    """셀 값 중 오류 토큰을 찾으면 위치 문자열을 반환(내용 수치는 미포함)."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.strip() in ERROR_TOKENS:
                        return f"{ws.title}!{cell.coordinate}"
    finally:
        wb.close()
    return None


def _scan_formulas_static(path: Path) -> str | None:
    """LibreOffice 가 없을 때: 수식 텍스트에 오류 토큰이 직접 박혀있는지 정적 점검."""
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and any(tok in v for tok in ERROR_TOKENS):
                        return f"{ws.title}!{cell.coordinate}"
    finally:
        wb.close()
    return None


def _recalc_with_libreoffice(src: Path, soffice: str) -> Path | None:
    """LibreOffice 로 재계산하여 새 xlsx 를 생성. 실패 시 None."""
    out_dir = Path(tempfile.mkdtemp(prefix="recalc_"))
    try:
        # 대용량 명세서 재계산도 끝까지 수행하도록 넉넉히(10분). Cloud Run 요청
        # 타임아웃(3600s) 안에서 명세서/지배구조 2회 재계산이 모두 들어간다.
        subprocess.run(
            [soffice, "--headless", "--calc",
             "--convert-to", "xlsx:Calc MS Excel 2007 XML",
             "--outdir", str(out_dir), str(src)],
            check=True, capture_output=True, timeout=600,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
        return None
    produced = out_dir / (src.stem + ".xlsx")
    return produced if produced.exists() else None


def verify_output(path: Path) -> CheckResult:
    # 1) openpyxl 재오픈
    try:
        wb = load_workbook(path, data_only=False)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        return CheckResult(False, f"출력 파일 재오픈 실패: {exc}", False)

    soffice = _find_soffice()
    if soffice:
        recalced = _recalc_with_libreoffice(path, soffice)
        if recalced is not None:
            bad = _scan_workbook_for_errors(recalced)
            shutil.rmtree(recalced.parent, ignore_errors=True)
            if bad:
                return CheckResult(False, f"재계산 후 수식 오류 발견: {bad}", True)
            return CheckResult(True, "재계산 후 수식 오류 0건", True)
        # 재계산 실패 → 정적 점검으로 폴백
    # LibreOffice 미설치/실패 → 정적 점검
    bad = _scan_formulas_static(path)
    if bad:
        return CheckResult(False, f"수식 오류 토큰 발견: {bad}", False)
    return CheckResult(True, "정적 점검 통과(LibreOffice 미사용)", False)
