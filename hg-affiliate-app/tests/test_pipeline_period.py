"""run_pipeline 의 당기 추출 연동 테스트 (외부 서비스 불필요)."""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.domain.errors import ErrorLog
from app.domain.period_extract import Period, parse_year_month
from app.domain.statement_detail import StatementDetailError
from app.output_check import CheckResult
from app.pipeline import run_pipeline

_SETTINGS = Settings(openai_api_key=None, openai_model="m")


def _make_ledger(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["계정명", "날짜", "거래처명", "차변", "대변"])
    for r in rows:
        ws.append(r)
    wb.save(path)


def _make_ledger_with_literal_description(path, description: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["계정명", "날짜", "거래처명", "적요", "차변", "대변"])
    ws.append(["상품매출", "2026-06-30", "특관자A", description, "0", "100"])
    description_cell = ws.cell(row=2, column=4)
    description_cell.data_type = "s"
    wb.save(path)
    wb.close()


def _slots(ledger=None):
    return {
        "prev_balance": None,
        "current_ledger": ledger,
        "current_balance": None,
        "name_pivot": None,
        "prev_statement": None,
    }


def _make_period_template(path):
    wb = Workbook()
    wb.active.title = "특관자"
    detail = wb.create_sheet("39.1")
    pivot = wb.create_sheet("39.2")
    detail["B15"] = "매입매출"
    detail["F15"] = "구분계정과목"
    detail["I15"] = "날짜"
    detail["J15"] = "적요"
    detail["J16"] = "OLD_Q1_SENTINEL"
    detail["P12"] = "=SUM(P16:P510)"
    pivot["D3"] = "특관자A"
    wb.save(path)


def test_run_pipeline_blocks_on_unparseable_ledger_date(tmp_path):
    # 날짜 해석 불가 셀 + AI 키 없음 → 엄격 차단(다운로드 차단)
    led = tmp_path / "ledger.xlsx"
    _make_ledger(led, [
        ["현금", "2026-01-10", "거래처A", "1000", "0"],
        ["현금", "날짜미상", "거래처B", "1000", "0"],
    ])
    outcome = run_pipeline(_slots(led), _SETTINGS, tmp_path, period=Period(2026, 1))
    assert outcome.ok is False
    assert "날짜" in outcome.message
    assert "오류목록" in outcome.outputs


def test_run_pipeline_no_block_when_dates_clean(tmp_path):
    # 모든 날짜 파싱 가능 → 추출 단계에서 차단되지 않음(이후 단계로 진행)
    led = tmp_path / "ledger.xlsx"
    _make_ledger(led, [
        ["현금", "2026-01-10", "거래처A", "1000", "0"],
        ["현금", "2026-02-20", "거래처B", "1000", "0"],
    ])
    outcome = run_pipeline(_slots(led), _SETTINGS, tmp_path, period=Period(2026, 1))
    # 날짜 차단 메시지가 아니어야 한다(다른 사유로 ok 여부는 무관)
    assert "날짜" not in outcome.message


def test_run_pipeline_q2_uses_january_through_june_for_summary_and_detail(
    tmp_path, monkeypatch
):
    from app import pipeline
    from app.domain import statement

    template = tmp_path / "template.xlsx"
    _make_period_template(template)
    monkeypatch.setattr(statement, "TEMPLATE_PATH", template)
    monkeypatch.setattr(
        pipeline,
        "verify_output",
        lambda _: CheckResult(ok=True, reason="test", recalculated=False),
    )
    seen: dict[str, object] = {}
    original_detail = pipeline.build_statement_detail
    original_aggregate = pipeline.aggregate_ledger
    original_reconstruction = pipeline._verify_reconstruction

    def capture_detail(ledger, **kwargs):
        seen["detail"] = ledger
        return original_detail(ledger, **kwargs)

    def capture_aggregate(ledger, *args, **kwargs):
        seen["aggregate"] = ledger
        return original_aggregate(ledger, *args, **kwargs)

    def capture_reconstruction(*args, **kwargs):
        seen["reconstruction"] = args[1]
        return original_reconstruction(*args, **kwargs)

    monkeypatch.setattr(pipeline, "build_statement_detail", capture_detail)
    monkeypatch.setattr(pipeline, "aggregate_ledger", capture_aggregate)
    monkeypatch.setattr(pipeline, "_verify_reconstruction", capture_reconstruction)
    ledger = tmp_path / "ledger.xlsx"
    _make_ledger(
        ledger,
        [
            ["상품매출", "2026-01-03", "특관자A", "0", "100"],
            ["상품매출", "2026-06-30", "특관자A", "0", "200"],
            ["상품매출", "2026-07-01", "특관자A", "0", "400"],
        ],
    )

    outcome = run_pipeline(
        _slots(ledger),
        _SETTINGS,
        tmp_path,
        period=Period(2026, 2),
    )

    assert outcome.ok is True
    assert seen["aggregate"] is seen["detail"] is seen["reconstruction"]
    assert seen["detail"].attrs["period_year_months"] == [(2026, 1), (2026, 6)]
    statement_path = outcome.outputs["당기_특관자_명세서"]
    assert statement_path.name == "당기_특관자_명세서_2026_2Q.xlsx"
    workbook = load_workbook(statement_path, data_only=False)
    detail = workbook["39.1"]
    assert [
        parse_year_month(detail["I16"].value),
        parse_year_month(detail["I17"].value),
    ] == [
        (2026, 1),
        (2026, 6),
    ]
    assert detail["I18"].value is None
    assert detail["A1"].value == "2026년 2분기 누적 (1~6월)"
    assert workbook["39.2"]["E3"].value == 300.0
    assert "OLD_Q1_SENTINEL" not in {
        detail.cell(row=row, column=10).value
        for row in range(16, detail.max_row + 1)
    }
    workbook.close()


def test_run_pipeline_avoids_iterrows_for_period_filtered_ledger(tmp_path, monkeypatch):
    from app import pipeline
    from app.domain import statement

    template = tmp_path / "template.xlsx"
    _make_period_template(template)
    monkeypatch.setattr(statement, "TEMPLATE_PATH", template)
    monkeypatch.setattr(
        pipeline,
        "verify_output",
        lambda _: CheckResult(ok=True, reason="test", recalculated=False),
    )

    def fail_if_iterrows_called(self):
        raise AssertionError("period-filtered ledger must not use DataFrame.iterrows()")

    monkeypatch.setattr(pd.DataFrame, "iterrows", fail_if_iterrows_called)
    ledger = tmp_path / "ledger.xlsx"
    _make_ledger(
        ledger,
        [["상품매출", "2026-06-30", "특관자A", "0", "100"]],
    )

    outcome = run_pipeline(
        _slots(ledger),
        _SETTINGS,
        tmp_path,
        period=Period(2026, 2),
    )

    assert outcome.ok is True


def test_reconstruction_avoids_iterrows_for_period_metadata(monkeypatch):
    from app import pipeline

    ledger = pd.DataFrame(
        [
            {
                "계정명": "외상매출금",
                "거래처명": "특관자A",
                "차변": "100",
                "대변": "0",
            }
        ]
    )
    prev_balance = pd.DataFrame(
        [{"계정과목": "외상매출금", "거래처명": "특관자A", "금액": "0"}]
    )
    current_balance = pd.DataFrame(
        [{"계정과목": "외상매출금", "거래처명": "특관자A", "금액": "100"}]
    )
    metadata = [(2026, 6)]
    attrs = ledger.attrs
    ledger.attrs["period_year_months"] = metadata

    def fail_if_iterrows_called(self):
        raise AssertionError("ledger reconstruction must not use DataFrame.iterrows()")

    monkeypatch.setattr(pd.DataFrame, "iterrows", fail_if_iterrows_called)
    errors = ErrorLog()

    pipeline._verify_reconstruction(
        prev_balance,
        ledger,
        current_balance,
        mapping={"특관자A": "특관자A"},
        canonical={"특관자A"},
        errors=errors,
    )

    assert ledger.attrs is attrs
    assert ledger.attrs["period_year_months"] is metadata
    assert errors.count == 0


def test_run_pipeline_q2_preserves_string_typed_error_looking_description_with_static_validation(
    tmp_path, monkeypatch
):
    from app import output_check, pipeline
    from app.domain import statement

    template = tmp_path / "template.xlsx"
    _make_period_template(template)
    monkeypatch.setattr(statement, "TEMPLATE_PATH", template)
    monkeypatch.setattr(output_check, "_find_soffice", lambda: None)
    monkeypatch.setattr(pipeline, "verify_output", output_check.verify_output)
    ledger = tmp_path / "ledger.xlsx"
    _make_ledger_with_literal_description(ledger, "#REF!")

    outcome = run_pipeline(
        _slots(ledger),
        _SETTINGS,
        tmp_path,
        period=Period(2026, 2),
    )

    assert outcome.ok is True
    statement_path = outcome.outputs["당기_특관자_명세서"]
    assert statement_path.name == "당기_특관자_명세서_2026_2Q.xlsx"
    workbook = load_workbook(statement_path, data_only=False)
    try:
        description = workbook["39.1"]["J16"]
        assert description.value == "#REF!"
        assert description.data_type == "s"
    finally:
        workbook.close()


def test_run_pipeline_blocks_detail_failure_without_leaking_accounting_values(
    tmp_path, monkeypatch
):
    from app import pipeline

    secret = "sensitive transaction description 12345"
    ledger = tmp_path / "ledger.xlsx"
    _make_ledger(ledger, [["상품매출", "2026-06-30", "특관자A", "0", "100"]])

    def fail_detail(*args, **kwargs):
        raise StatementDetailError(secret)

    monkeypatch.setattr(pipeline, "build_statement_detail", fail_detail)

    outcome = run_pipeline(_slots(ledger), _SETTINGS, tmp_path, period=Period(2026, 2))

    assert outcome.ok is False
    assert outcome.message == "당기 상세 거래 검증에 실패하여 다운로드를 차단했습니다."
    assert set(outcome.outputs) == {"오류목록"}
    error_book = load_workbook(outcome.outputs["오류목록"], data_only=True)
    error_values = {
        str(cell.value)
        for row in error_book.active.iter_rows()
        for cell in row
        if cell.value is not None
    }
    error_book.close()
    assert "당기 상세 검증 실패" in error_values
    assert secret not in error_values


@pytest.mark.parametrize("failure_stage", ["build", "save"])
def test_run_pipeline_statement_writer_failure_is_generic_and_never_leaks_exception(
    tmp_path, monkeypatch, failure_stage
):
    from app import pipeline

    sentinel = "TRANSACTION_SENTINEL_SHOULD_NEVER_ESCAPE"
    ledger = tmp_path / "ledger.xlsx"
    _make_ledger(ledger, [["상품매출", "2026-06-30", "특관자A", "0", "100"]])

    if failure_stage == "build":
        def fail_build(*args, **kwargs):
            raise RuntimeError(sentinel)

        monkeypatch.setattr(pipeline, "build_statement", fail_build)
    else:
        class FailingWorkbook:
            def save(self, path):
                raise RuntimeError(sentinel)

        monkeypatch.setattr(
            pipeline,
            "build_statement",
            lambda *args, **kwargs: (FailingWorkbook(), []),
        )

    outcome = run_pipeline(_slots(ledger), _SETTINGS, tmp_path, period=Period(2026, 2))

    assert outcome.ok is False
    assert outcome.message == "당기 특관자 명세서 생성에 실패하여 다운로드를 차단했습니다."
    assert sentinel not in outcome.message
    assert set(outcome.outputs) == {"오류목록"}
    error_book = load_workbook(outcome.outputs["오류목록"], data_only=True)
    error_values = [
        str(cell.value)
        for row in error_book.active.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    error_book.close()
    assert all(sentinel not in value for value in error_values)
