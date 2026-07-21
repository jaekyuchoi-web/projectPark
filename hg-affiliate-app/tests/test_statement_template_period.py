from __future__ import annotations

from pathlib import Path
from zipfile import BadZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.domain.aggregate import AggregateResult
from app.domain.period_extract import Period
from app.domain.statement_detail import StatementDetailError, StatementDetailRow
from app.domain.statement_template import (
    StatementTemplateError,
    fill_statement_template,
)


def _write_template(path: Path) -> None:
    wb = Workbook()
    wb.active.title = "특관자"
    detail = wb.create_sheet("39.1")
    pivot = wb.create_sheet("39.2")
    detail["B15"] = "매입매출"
    detail["F15"] = "구분계정과목"
    detail["I15"] = "날짜"
    detail["J15"] = "적요"
    detail["B16"] = "OLD_Q1_SENTINEL"
    detail["J16"] = "OLD_Q1_SENTINEL"
    detail["B16"].fill = PatternFill("solid", fgColor="FFFF00")
    detail["C4"] = "=SUMIFS($N$16:$N$510,$F$16:$F$510,$B4)"
    detail["P12"] = "=SUM(P16:P510)"
    pivot["D3"] = "특관자A"
    wb.save(path)


def _row(date: str, description: str) -> StatementDetailRow:
    return StatementDetailRow(
        sales_purchase="매출",
        funding=None,
        receivable_payable=None,
        income_expense=None,
        bucket="매출",
        account_code="40100",
        account_name="상품매출",
        date=date,
        description=description,
        partner_code="V001",
        partner_name="특관자A",
        canonical_name="특관자A",
        debit=0.0,
        credit=100.0,
        balance=100.0,
    )


def _set_literal_text(cell, value: str) -> None:
    cell.value = value
    cell.data_type = "s"


def _assert_period_label(ws) -> None:
    assert ws["A1"].value == "2026년 2분기 누적 (1~6월)"
    assert "A1:F1" in {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A1"].alignment.vertical == "center"
    assert ws["A1"].font.bold is True


def test_template_replaces_stale_q1_detail_with_q2_ytd_rows(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    source = load_workbook(template)
    source["39.1"]["A2"] = "ROW_TWO_SENTINEL"
    source.save(template)
    source.close()

    wb, unmatched = fill_statement_template(
        template,
        AggregateResult(),
        Period(2026, 2),
        [_row("2026-01-03", "2026.1Q 정상 적요"), _row("2026-06-30", "June")],
    )

    detail = wb["39.1"]
    assert unmatched == []
    _assert_period_label(detail)
    assert detail["A2"].value == "ROW_TWO_SENTINEL"
    assert detail["I16"].value == "2026-01-03"
    assert detail["I17"].value == "2026-06-30"
    assert detail["J16"].value == "2026.1Q 정상 적요"
    assert "OLD_Q1_SENTINEL" not in {
        detail.cell(row=row, column=column).value
        for row in range(16, detail.max_row + 1)
        for column in range(2, 17)
    }
    assert detail["C4"].value == "=SUMIFS($N$16:$N$17,$F$16:$F$17,$B4)"
    assert detail["P12"].value == "=SUM(P16:P17)"
    assert detail["B17"].fill.fgColor.rgb == detail["B16"].fill.fgColor.rgb


def test_template_writer_preserves_equals_prefixed_detail_as_literal_text(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    _write_template(template)

    wb, _ = fill_statement_template(
        template,
        AggregateResult(),
        Period(2026, 2),
        [_row("2026-06-30", "=1+1")],
    )
    wb.save(output)
    wb.close()

    loaded = load_workbook(output, data_only=False)
    cell = loaded["39.1"]["J16"]
    assert cell.value == "=1+1"
    assert cell.data_type == "s"
    loaded.close()


def test_template_writer_preserves_formula_and_error_looking_detail_text(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    _write_template(template)
    literals = ["=[29] literal", "=#REF! literal", "#REF!"]

    wb, _ = fill_statement_template(
        template,
        AggregateResult(),
        Period(2026, 2),
        [_row("2026-06-30", literal) for literal in literals],
    )
    wb.save(output)
    wb.close()

    loaded = load_workbook(output, data_only=False)
    for row, literal in enumerate(literals, start=16):
        cell = loaded["39.1"].cell(row=row, column=10)
        assert cell.value == literal
        assert cell.data_type == "s"
    loaded.close()


def test_summary_rewrite_ignores_string_typed_formula_looking_top_cell(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    wb = load_workbook(template)
    _set_literal_text(wb["39.1"]["P12"], "=SUM(P16:P510)")
    wb.save(template)
    wb.close()

    output, _ = fill_statement_template(
        template,
        AggregateResult(),
        Period(2026, 2),
        [_row("2026-06-30", "June")],
    )

    literal = output["39.1"]["P12"]
    assert literal.value == "=SUM(P16:P510)"
    assert literal.data_type == "s"
    assert output["39.1"]["C4"].value == "=SUMIFS($N$16:$N$16,$F$16:$F$16,$B4)"


def test_string_typed_formula_looking_top_cell_does_not_satisfy_summary_requirement(
    tmp_path,
):
    template = tmp_path / "template.xlsx"
    wb = Workbook()
    wb.active.title = "특관자"
    detail = wb.create_sheet("39.1")
    wb.create_sheet("39.2")
    _set_literal_text(detail["P12"], "=SUM(P16:P510)")
    wb.save(template)
    wb.close()

    with pytest.raises(ValueError) as exc_info:
        fill_statement_template(template, AggregateResult(), Period(2026, 2), [])

    assert str(exc_info.value) == "39.1 요약 수식을 안전하게 갱신하지 못했습니다."


def test_template_cleanup_neutralizes_actual_formula_and_error_cells(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    wb = load_workbook(template)
    pivot = wb["39.2"]
    pivot["L3"] = "=[29]external.xlsx!A1"
    pivot["M3"] = "#REF!"
    assert pivot["L3"].data_type == "f"
    assert pivot["M3"].data_type == "e"
    wb.save(template)
    wb.close()

    output, _ = fill_statement_template(
        template, AggregateResult(), Period(2026, 2), []
    )

    assert output["39.2"]["L3"].value == 0
    assert output["39.2"]["M3"].value is None


def test_legacy_writer_preserves_equals_prefixed_detail_as_literal_text(
    tmp_path, monkeypatch
):
    from app.domain import statement

    monkeypatch.setattr(statement, "TEMPLATE_PATH", tmp_path / "missing.xlsx")
    output = tmp_path / "legacy.xlsx"
    wb, _ = statement.build_statement(
        AggregateResult(),
        Period(2026, 2),
        [_row("2026-06-30", "=1+1")],
    )
    wb.save(output)
    wb.close()

    loaded = load_workbook(output, data_only=False)
    cell = loaded["39.1"]["J16"]
    assert cell.value == "=1+1"
    assert cell.data_type == "s"
    loaded.close()


def test_legacy_writer_merges_and_centers_period_label_without_changing_detail(
    tmp_path, monkeypatch
):
    from app.domain import statement

    monkeypatch.setattr(statement, "TEMPLATE_PATH", tmp_path / "missing.xlsx")
    wb, _ = statement.build_statement(
        AggregateResult(),
        Period(2026, 2),
        [_row("2026-06-30", "June")],
    )

    detail = wb["39.1"]
    _assert_period_label(detail)
    assert detail["B15"].value == "매입매출"
    assert detail["I16"].value == "2026-06-30"
    assert detail["J16"].value == "June"


def test_template_period_label_accepts_the_exact_existing_merge(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    source = load_workbook(template)
    source["39.1"].merge_cells("A1:F1")
    source.save(template)
    source.close()

    wb, _ = fill_statement_template(
        template, AggregateResult(), Period(2026, 2), []
    )

    _assert_period_label(wb["39.1"])
    assert len(wb["39.1"].merged_cells.ranges) == 1


def test_template_period_label_rejects_an_incompatible_overlapping_merge(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    source = load_workbook(template)
    source["39.1"].merge_cells("B1:G1")
    source.save(template)
    source.close()

    with pytest.raises(StatementTemplateError) as exc_info:
        fill_statement_template(template, AggregateResult(), Period(2026, 2), [])

    assert str(exc_info.value) == "39.1 기간 라벨 영역을 안전하게 갱신하지 못했습니다."
    assert "B1:G1" not in str(exc_info.value)


def test_illegal_detail_text_raises_generic_statement_detail_error(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    secret = "PRIVATE_TRANSACTION\x01SENTINEL"

    with pytest.raises(StatementDetailError) as exc_info:
        fill_statement_template(
            template,
            AggregateResult(),
            Period(2026, 2),
            [_row("2026-06-30", secret)],
        )

    assert str(exc_info.value) == "39.1 상세 거래를 안전하게 기록하지 못했습니다."
    assert "PRIVATE_TRANSACTION" not in str(exc_info.value)


def test_formula_rewrite_changes_only_local_or_self_qualified_detail_ranges(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    wb = load_workbook(template)
    detail = wb["39.1"]
    detail["D4"] = (
        '=SUM(N16:N510)+SUM(\'39.1\'!$P$16:$P$510)'
        '+SUM(\'39.2\'!B16:B44)+SUM(BA16:BA510)+"N16:N510"'
    )
    wb.save(template)
    wb.close()

    output, _ = fill_statement_template(
        template,
        AggregateResult(),
        Period(2026, 2),
        [_row("2026-01-03", "January"), _row("2026-06-30", "June")],
    )

    assert output["39.1"]["D4"].value == (
        '=SUM(N16:N17)+SUM(\'39.1\'!$P$16:$P$17)'
        '+SUM(\'39.2\'!B16:B44)+SUM(BA16:BA510)+"N16:N510"'
    )


def test_template_without_recognized_local_detail_summary_range_fails_closed(tmp_path):
    template = tmp_path / "template.xlsx"
    wb = Workbook()
    wb.active.title = "특관자"
    detail = wb.create_sheet("39.1")
    wb.create_sheet("39.2")
    detail["C4"] = "=SUM('39.2'!N16:N510)"
    wb.save(template)
    wb.close()

    with pytest.raises(ValueError) as exc_info:
        fill_statement_template(template, AggregateResult(), Period(2026, 2), [])

    assert str(exc_info.value) == "39.1 요약 수식을 안전하게 갱신하지 못했습니다."
    assert "N16:N510" not in str(exc_info.value)


def test_zero_detail_clears_stale_values_and_sets_local_endpoint_to_row_16(tmp_path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    wb = load_workbook(template)
    detail = wb["39.1"]
    detail["D4"] = (
        '=SUM(N16:N510)+SUM(\'39.2\'!B16:B44)+"P16:P510"'
    )
    detail["P40"] = "STALE_DETAIL"
    wb.save(template)
    wb.close()

    output, _ = fill_statement_template(
        template, AggregateResult(), Period(2026, 2), []
    )
    detail = output["39.1"]

    assert all(
        detail.cell(row=row, column=column).value is None
        for row in range(16, detail.max_row + 1)
        for column in range(2, 17)
    )
    assert detail["C4"].value == "=SUMIFS($N$16:$N$16,$F$16:$F$16,$B4)"
    assert detail["P12"].value == "=SUM(P16:P16)"
    assert detail["D4"].value == (
        '=SUM(N16:N16)+SUM(\'39.2\'!B16:B44)+"P16:P510"'
    )


def test_existing_broken_template_is_not_silently_hidden_by_fallback(tmp_path, monkeypatch):
    from app.domain import statement

    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"not an xlsx")
    monkeypatch.setattr(statement, "TEMPLATE_PATH", broken)

    with pytest.raises(BadZipFile):
        statement.build_statement(AggregateResult(), Period(2026, 2), [])
