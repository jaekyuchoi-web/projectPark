from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app import output_check


def _set_string(cell, value: str) -> None:
    cell.value = value
    cell.data_type = "s"


def _workbook(path: Path, setup) -> None:
    wb = Workbook()
    setup(wb.active)
    wb.save(path)
    wb.close()


def test_literal_error_looking_strings_pass_static_and_recalculated_scans(tmp_path):
    path = tmp_path / "literal.xlsx"

    def setup(ws) -> None:
        _set_string(ws["A1"], "#REF!")
        _set_string(ws["A2"], "=#REF! literal")

    _workbook(path, setup)

    loaded = load_workbook(path, data_only=False, read_only=True)
    try:
        assert loaded.active["A1"].data_type == "s"
        assert loaded.active["A2"].data_type == "s"
    finally:
        loaded.close()

    assert output_check._scan_formulas_static(path) is None
    assert output_check._scan_workbook_for_errors(path) is None


def test_verify_output_static_fallback_accepts_literal_error_looking_strings(
    tmp_path, monkeypatch
):
    path = tmp_path / "literal.xlsx"
    _workbook(path, lambda ws: _set_string(ws["A1"], "#REF!"))
    monkeypatch.setattr(output_check, "_find_soffice", lambda: None)

    result = output_check.verify_output(path)

    assert result.ok is True
    assert result.recalculated is False


def test_actual_excel_error_cell_fails_static_and_recalculated_scans(tmp_path):
    path = tmp_path / "error.xlsx"
    _workbook(path, lambda ws: ws.__setitem__("A1", "#REF!"))

    loaded = load_workbook(path, data_only=False, read_only=True)
    try:
        assert loaded.active["A1"].data_type == "e"
    finally:
        loaded.close()

    assert output_check._scan_formulas_static(path) == "Sheet!A1"
    assert output_check._scan_workbook_for_errors(path) == "Sheet!A1"


def test_actual_formula_with_error_token_fails_static_scan(tmp_path):
    path = tmp_path / "formula.xlsx"
    _workbook(path, lambda ws: ws.__setitem__("A1", "=IFERROR(#REF!, 0)"))

    loaded = load_workbook(path, data_only=False, read_only=True)
    try:
        assert loaded.active["A1"].data_type == "f"
    finally:
        loaded.close()

    assert output_check._scan_formulas_static(path) == "Sheet!A1"
